#!/usr/bin/env python3
"""Exporta os cultos da planilha `BOLETIM.xlsm` para JSON e resuma o conjunto."""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterator, List

DATE_PATTERN = re.compile(
    r"(\d{1,2})\s+DE\s+([A-Za-zÀ-ÿ]+)\s+DE\s+(\d{4})", re.IGNORECASE
)
MONTHS = {
    "JANEIRO": 1,
    "FEVEREIRO": 2,
    "MARÇO": 3,
    "MARCO": 3,
    "ABRIL": 4,
    "MAIO": 5,
    "JUNHO": 6,
    "JULHO": 7,
    "AGOSTO": 8,
    "SETEMBRO": 9,
    "OUTUBRO": 10,
    "NOVEMBRO": 11,
    "DEZEMBRO": 12,
}


def parse_portuguese_date(value: str) -> date | None:
    if not value:
        return None
    value = value.strip().upper()
    match = DATE_PATTERN.search(value)
    if not match:
        return None
    day, month, year = match.groups()
    month_number = MONTHS.get(month)
    if month_number is None:
        return None
    try:
        return date(year=int(year), month=month_number, day=int(day))
    except ValueError:
        return None


def normalize_header(header: str | None, seen: Counter[str]) -> str:
    base = header or "CAMPO"
    base = unicodedata.normalize("NFKD", base)
    base = "".join(ch for ch in base if not unicodedata.category(ch) == "Mn")
    base = re.sub(r"\s+", "_", base)
    base = re.sub(r"[^\w]", "_", base)
    base = base.strip("_").upper() or "CAMPO"
    count = seen[base]
    seen[base] += 1
    return f"{base}_{count+1}" if count else base


def iter_rows(sheet) -> Iterator[Dict[str, Any]]:
    headers = [cell.value for cell in sheet[1]]
    seen = Counter[str]()
    keys = [normalize_header(h, seen) for h in headers]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        yield dict(zip(keys, row))


def build_export(path: Path) -> Dict[str, Any]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    if "CULTOS" not in wb.sheetnames:
        raise ValueError("A planilha CULTOS não está presente.")
    sheet = wb["CULTOS"]
    rows = []
    counts_by_year: Counter[int] = Counter()
    counts_by_dirigente: Counter[str] = Counter()
    earliest: date | None = None
    latest: date | None = None
    for record in iter_rows(sheet):
        date_text = record.get("DATA")
        parsed = parse_portuguese_date(str(date_text)) if date_text else None
        iso = parsed.isoformat() if parsed else None
        if parsed:
            counts_by_year[parsed.year] += 1
            earliest = parsed if earliest is None else min(earliest, parsed)
            latest = parsed if latest is None else max(latest, parsed)
        dirigente = record.get("DIRIGENTE")
        if dirigente:
            counts_by_dirigente[dirigente] += 1
        record_clean = {
            "date_text": date_text,
            "date_iso": iso,
            "dirigente": dirigente,
            "raw": record,
        }
        rows.append(record_clean)
    summary = {
        "total_cultos": len(rows),
        "with_dates": sum(1 for r in rows if r["date_iso"]),
        "without_dates": sum(1 for r in rows if not r["date_iso"]),
        "earliest": earliest.isoformat() if earliest else None,
        "latest": latest.isoformat() if latest else None,
        "by_year": dict(sorted(counts_by_year.items())),
        "by_dirigente": counts_by_dirigente.most_common(10),
    }
    return {"cultos": rows, "summary": summary}


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extrai os dados de cultos de BOLETIM.xlsm para JSON."
    )
    parser.add_argument(
        "--source",
        "-s",
        default="BOLETIM.xlsm",
        help="Arquivo .xlsm (padrão: BOLETIM.xlsm)",
    )
    parser.add_argument(
        "--output",
        "-o",
        default="data/cultos.json",
        help="JSON de saída (criando pastas se necessário).",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_arguments()
    source = Path(args.source)
    destination = Path(args.output)
    if not source.exists():
        raise SystemExit(f"{source} não encontrado.")
    export = build_export(source)
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_text(
        json.dumps(export, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    summary = export["summary"]
    print("Exportação concluída.")
    print(f"  cultos exportados: {summary['total_cultos']}")
    print(f"  com datas:        {summary['with_dates']} / sem data: {summary['without_dates']}")
    print(f"  intervalo:        {summary['earliest']} -> {summary['latest']}")
    print("  cultos por ano:", summary["by_year"])
    print("  top dirigentes:", ", ".join(f"{name} ({count})" for name, count in summary["by_dirigente"]))


if __name__ == "__main__":
    main()
