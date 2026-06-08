from __future__ import annotations

import json
import re
import zipfile
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook

from src.config import LEGACY_DIR, LEGACY_XLSM_PATH
from src.database import Database
from src.logger import get_logger
from src.models import ImportResult, WorshipBulletin


logger = get_logger(__name__)


LEGACY_COLUMN_MAP = {
    1: "date_text",
    2: "dirigente",
    3: "preludio_musica",
    4: "preludio_cantor",
    5: "preludio_tom",
    6: "ref1",
    7: "texto1",
    8: "musica1",
    9: "cantor1",
    10: "tom1",
    11: "ref2",
    12: "texto2",
    13: "musica2",
    14: "cantor2",
    15: "tom2",
    16: "ref3",
    17: "texto3",
    18: "musica3",
    19: "cantor3",
    20: "tom3",
    21: "oracao_louvor",
    22: "ref_louvor",
    23: "texto_louvor",
    24: "ofertas_ref",
    25: "ofertas_texto",
    26: "ofertas_oracao",
    27: "musica4",
    28: "cantor4",
    29: "tom4",
    30: "musica5",
    31: "cantor5",
    32: "tom5",
    33: "oracao_intercessao",
    34: "pregador",
    35: "musica_pao",
    36: "cantor_pao",
    37: "tom_pao",
    38: "musica_vinho",
    39: "cantor_vinho",
    40: "tom_vinho",
    41: "musica_extra",
    42: "cantor_extra",
    43: "tom_extra",
    44: "musica_final",
    45: "cantor_final",
    46: "tom_final",
}


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    return str(value).strip()


def find_legacy_file() -> Path:
    if LEGACY_XLSM_PATH.exists():
        return LEGACY_XLSM_PATH
    candidates = [
        path
        for path in sorted(LEGACY_DIR.glob("*.xlsm"))
        if not path.name.lower().startswith("coloque_aqui")
    ]
    if candidates:
        return candidates[0]
    raise FileNotFoundError(
        "Arquivo XLSM nao encontrado. Coloque BOLETIM_VBA_CORRIGIDO.xlsm na pasta legacy/."
    )


def inspect_legacy_workbook(workbook_path: Optional[str | Path] = None) -> dict[str, Any]:
    path = Path(workbook_path) if workbook_path else find_legacy_file()
    if not path.exists():
        raise FileNotFoundError(f"Arquivo XLSM nao encontrado: {path}")

    try:
        workbook = load_workbook(path, keep_vba=True, data_only=True, read_only=True)
        try:
            sheets = workbook.sheetnames
            has_planilha1 = "Planilha1" in sheets
            has_louvor = "LOUVOR" in sheets
            headers: list[str] = []
            max_row = 0
            max_column = 0
            if has_planilha1:
                sheet = workbook["Planilha1"]
                max_row = sheet.max_row
                max_column = sheet.max_column
                headers = [_cell_to_text(sheet.cell(row=1, column=column).value) for column in range(1, 47)]
            summary = {
                "path": str(path),
                "sheets": sheets,
                "has_planilha1": has_planilha1,
                "has_louvor": has_louvor,
                "planilha1_range": f"A1:AT{max_row}" if has_planilha1 else "",
                "planilha1_max_row": max_row,
                "planilha1_max_column": max_column,
                "headers": headers,
                "vba_summary": try_extract_vba_summary(path),
            }
        finally:
            workbook.close()
        logger.info("XLSM inspecionado: %s", path)
        return summary
    except Exception:
        logger.exception("Falha ao inspecionar XLSM legado")
        raise


def import_legacy_bulletins(workbook_path: Optional[str | Path] = None, db: Optional[Database] = None) -> ImportResult:
    path = Path(workbook_path) if workbook_path else find_legacy_file()
    database = db or Database()
    imported = 0
    skipped = 0
    errors: list[str] = []

    try:
        workbook = load_workbook(path, keep_vba=True, data_only=True, read_only=True)
        try:
            if "Planilha1" not in workbook.sheetnames:
                raise ValueError("Aba Planilha1 nao encontrada no arquivo XLSM.")
            sheet = workbook["Planilha1"]
            for row_index in range(2, sheet.max_row + 1):
                values = {
                    field_name: _cell_to_text(sheet.cell(row=row_index, column=column_index).value)
                    for column_index, field_name in LEGACY_COLUMN_MAP.items()
                }
                if not values["date_text"]:
                    continue

                legacy_values = {
                    str(column_index): _cell_to_text(sheet.cell(row=row_index, column=column_index).value)
                    for column_index in range(1, 47)
                }
                raw_json = json.dumps(
                    {"legacy_file": path.name, "legacy_row": row_index, "values": legacy_values},
                    ensure_ascii=False,
                )
                if database.bulletin_exists("legacy_xlsm", values["date_text"], raw_json):
                    skipped += 1
                    continue
                bulletin = WorshipBulletin(**values, source="legacy_xlsm", raw_json=raw_json)
                database.insert_bulletin(bulletin)
                imported += 1
        finally:
            workbook.close()

        message = f"{imported} boletim(ns) importado(s), {skipped} duplicata(s) ignorada(s)."
        database.insert_import_log("legacy_xlsm", "success", message)
        logger.info(message)
        return ImportResult("legacy_xlsm", "success", message, imported, skipped, errors)
    except Exception as exc:
        logger.exception("Falha ao importar boletins do XLSM legado")
        message = f"Falha ao importar boletins do XLSM: {exc}"
        database.insert_import_log("legacy_xlsm", "error", message)
        return ImportResult("legacy_xlsm", "error", message, imported, skipped, [str(exc)])


def import_legacy_people(workbook_path: Optional[str | Path] = None, db: Optional[Database] = None) -> ImportResult:
    path = Path(workbook_path) if workbook_path else find_legacy_file()
    database = db or Database()
    imported = 0
    skipped = 0
    errors: list[str] = []

    try:
        workbook = load_workbook(path, keep_vba=True, data_only=True, read_only=True)
        try:
            if "LOUVOR" not in workbook.sheetnames:
                raise ValueError("Aba LOUVOR nao encontrada no arquivo XLSM.")
            sheet = workbook["LOUVOR"]
            existing = {person.name.casefold() for person in database.list_people()}
            for row_index in range(2, sheet.max_row + 1):
                name = _cell_to_text(sheet.cell(row=row_index, column=1).value)
                if not name:
                    continue
                if name.casefold() in existing:
                    skipped += 1
                    continue
                database.save_person(name)
                existing.add(name.casefold())
                imported += 1
        finally:
            workbook.close()

        message = f"{imported} pessoa(s) importada(s), {skipped} duplicata(s) ignorada(s)."
        database.insert_import_log("legacy_xlsm", "success", message)
        logger.info(message)
        return ImportResult("legacy_xlsm", "success", message, imported, skipped, errors)
    except Exception as exc:
        logger.exception("Falha ao importar pessoas do XLSM legado")
        message = f"Falha ao importar pessoas do louvor: {exc}"
        database.insert_import_log("legacy_xlsm", "error", message)
        return ImportResult("legacy_xlsm", "error", message, imported, skipped, [str(exc)])


def try_extract_vba_summary(workbook_path: Optional[str | Path] = None) -> dict[str, Any]:
    path = Path(workbook_path) if workbook_path else find_legacy_file()
    summary = {"has_vba_project": False, "macro_like_names": [], "strings": []}
    try:
        with zipfile.ZipFile(path) as archive:
            names = archive.namelist()
            if "xl/vbaProject.bin" not in names:
                return summary
            summary["has_vba_project"] = True
            data = archive.read("xl/vbaProject.bin")
    except Exception as exc:
        summary["error"] = str(exc)
        return summary

    ascii_strings = [
        item.decode("latin1", errors="ignore")
        for item in re.findall(rb"[A-Za-z0-9_ .,:;!?()/\\-]{5,}", data)
    ]
    interesting = []
    macro_pattern = re.compile(r"\b(?:Sub\s+)?[A-Za-z0-9_]+_Click\b|\bbtn[A-Za-z0-9_]+\b", re.IGNORECASE)
    for item in ascii_strings:
        if macro_pattern.search(item):
            interesting.append(item.strip())

    summary["macro_like_names"] = sorted(set(interesting))[:50]
    summary["strings"] = sorted(set(ascii_strings))[:100]
    return summary
