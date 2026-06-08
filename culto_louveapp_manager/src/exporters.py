from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

from src.config import OUTPUT_DIR
from src.database import Database
from src.logger import get_logger
from src.models import BULLETIN_FIELDS, PERSON_FIELDS, SCHEDULE_FIELDS


logger = get_logger(__name__)


def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _write_workbook(path: Path, sheets: dict[str, pd.DataFrame]) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, dataframe in sheets.items():
            dataframe.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
            worksheet = writer.sheets[sheet_name]
            worksheet["A1"] = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
            worksheet["A1"].font = Font(bold=True)
            header_row = 3
            worksheet.freeze_panes = "A4"
            if dataframe.columns.size:
                last_column = get_column_letter(len(dataframe.columns))
                last_row = max(header_row, header_row + len(dataframe))
                worksheet.auto_filter.ref = f"A{header_row}:{last_column}{last_row}"
            for cell in worksheet[header_row]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="2F5597")
            for column_cells in worksheet.columns:
                max_length = 10
                column_letter = column_cells[0].column_letter
                for cell in column_cells:
                    value = "" if cell.value is None else str(cell.value)
                    max_length = max(max_length, min(len(value), 60))
                worksheet.column_dimensions[column_letter].width = max_length + 2
    logger.info("Arquivo Excel exportado: %s", path)
    return path


def export_bulletins_xlsx(output_path: str | Path | None = None) -> Path:
    database = Database()
    bulletins = pd.DataFrame(database.fetch_table_dicts("worship_bulletins"), columns=BULLETIN_FIELDS)
    people = pd.DataFrame(database.fetch_table_dicts("worship_people"), columns=PERSON_FIELDS)
    schedules = pd.DataFrame(database.fetch_table_dicts("louveapp_schedules"), columns=SCHEDULE_FIELDS)
    path = Path(output_path) if output_path else OUTPUT_DIR / f"boletins_{_timestamp()}.xlsx"
    return _write_workbook(
        path,
        {
            "Boletins": bulletins,
            "Pessoas": people,
            "Escalas LouveApp": schedules,
        },
    )


def export_bulletins_csv(output_path: str | Path | None = None) -> Path:
    database = Database()
    bulletins = pd.DataFrame(database.fetch_table_dicts("worship_bulletins"), columns=BULLETIN_FIELDS)
    path = Path(output_path) if output_path else OUTPUT_DIR / f"boletins_{_timestamp()}.csv"
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    bulletins.to_csv(path, index=False, encoding="utf-8-sig")
    logger.info("Boletins exportados para CSV: %s", path)
    return path


def export_louveapp_schedules_xlsx(output_path: str | Path | None = None) -> Path:
    database = Database()
    schedules = pd.DataFrame(database.fetch_table_dicts("louveapp_schedules"), columns=SCHEDULE_FIELDS)
    people = pd.DataFrame(database.fetch_table_dicts("worship_people"), columns=PERSON_FIELDS)
    path = Path(output_path) if output_path else OUTPUT_DIR / f"escalas_louveapp_{_timestamp()}.xlsx"
    return _write_workbook(
        path,
        {
            "Escalas LouveApp": schedules,
            "Pessoas": people,
        },
    )


def export_louveapp_schedules_csv(output_path: str | Path | None = None) -> Path:
    database = Database()
    schedules = pd.DataFrame(database.fetch_table_dicts("louveapp_schedules"), columns=SCHEDULE_FIELDS)
    path = Path(output_path) if output_path else OUTPUT_DIR / f"escalas_louveapp_{_timestamp()}.csv"
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    schedules.to_csv(path, index=False, encoding="utf-8-sig")
    logger.info("Escalas LouveApp exportadas para CSV: %s", path)
    return path
